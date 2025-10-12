import os
import sys
import cv2
import dlib
import base64
import pickle
import pymongo
import smtplib
import numpy as np
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file
from email.mime.text import MIMEText
from openpyxl import Workbook, load_workbook

# ==================== Flask Setup ====================
app = Flask(__name__)
app.secret_key = 'your_secret_key'

# ==================== MongoDB (Local) ====================
try:
    client = pymongo.MongoClient("mongodb://localhost:27017/")
    db = client["attendance_system"]
    attendance_collection = db["Attendance"]
    print("✅ Connected to MongoDB (localhost)")
except Exception as e:
    print(f"❌ MongoDB connection failed: {e}")
    sys.exit()

EXCEL_FILE = "attendance_records.xlsx"

# ==================== Face Recognition Setup ====================
detector = dlib.get_frontal_face_detector()
predictor = dlib.shape_predictor("models/shape_predictor_68_face_landmarks.dat")
face_rec_model = dlib.face_recognition_model_v1("models/dlib_face_recognition_resnet_model_v1.dat")

# Load known encodings
try:
    with open("face_encodings.pkl", "rb") as f:
        face_encodings = pickle.load(f)
except FileNotFoundError:
    print("❌ face_encodings.pkl not found.")
    sys.exit()

# ==================== Email Setup ====================
EMAIL = "automatedattendancemarker@gmail.com"
PASSWORD = "xnvj gcbl gqak gzbf"
people_emails = {
    "varsh": "varshathgannaram963@gmail.com",
    "pavan": "pavanrakhi229@gmail.com",
    "Rajeev": "kothapallyrajeev.r@gmail.com",
    "rakesh": "kothapallyrakesh.r@gmail.com",
    "bhanu teja": "24r05a6602@cmrithyderabad.edu.in",
}
emails_sent = set()

# ==================== Timetable ====================
timetable = {
    'Monday': {"09:10": "CRT"},
    'Tuesday': {
        "09:10": "OS - Bhavani",
        "10:10": "FSD - Rampriya",
        "11:10": "Break",
        "11:20": "DMGT - Sravanthi",
        "12:20": "DAA - Jasmine",
        "13:20": "LUNCH",
        "14:05": "RTSR - Sathish"
    },
    'Wednesday': {
        "09:10": "NODE JS LAB - Rampriya",
        "11:10": "Break",
        "11:20": "DAA - Jasmine",
        "12:20": "DMGT - Sravanthi",
        "13:20": "LUNCH",
        "14:05": "OS - Bhavani",
        "15:05": "ICC - Kolishetty"
    },
    'Thursday': {
        "09:10": "FSD - Rampriya",
        "10:10": "CN - Sathish kumar",
        "11:10": "Break",
        "11:20": "RTSR - Sathish",
        "13:20": "LUNCH",
        "14:05": "OS - Bhavani"
    },
    'Friday': {
        "09:10": "DMGT - Sravanthi",
        "10:10": "DAA - Jasmine",
        "11:10": "Break",
        "11:20": "CN - Sathish kumar",
        "12:20": "FSD - Rampriya",
        "13:20": "LUNCH",
        "14:05": "ATT - Mounika",
    },
    'Saturday': {
        "09:10": "CN - Sathish kumar",
        "10:10": "FSD - Rampriya",
        "11:10": "Break",
        "11:20": "CN & OS - Sathish kumar",
        "13:20": "LUNCH",
        "14:05": "ICC - Kolishetty",
        "15:05": "ES-II - Thirupathi Reddy"
    }
}

# ==================== Helpers ====================
def get_current_class():
    """Return current class name & time from timetable."""
    day = datetime.today().strftime('%A')
    current_time = datetime.now().strftime('%H:%M')
    schedule = timetable.get(day, {})
    if not schedule:
        return "No class today", "Unknown"
    class_time = max((t for t in schedule if t <= current_time), default=None)
    return schedule.get(class_time, "No class at this time"), class_time or "Unknown"

def log_attendance(name, class_name, class_time):
    """Log attendance in MongoDB + Excel with duplicate check."""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    current_date = datetime.now().strftime('%Y-%m-%d')

    # Mongo duplicate check
    if attendance_collection.find_one({"name": name, "class_name": class_name, "date": current_date}):
        print(f"⚠️ Duplicate attendance skipped for {name}")
        return

    # Save in Mongo
    attendance_collection.insert_one({
        "timestamp": timestamp,
        "name": name,
        "class_name": class_name,
        "class_time": class_time,
        "date": current_date
    })

    # Save in Excel
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"
            ws.append(["Name", "Class", "Class Time", "Timestamp", "Date"])
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Attendance"]

        # Prevent duplicates in Excel
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == name and row[1] == class_name and row[4] == current_date:
                return

        ws.append([name, class_name, class_time, timestamp, current_date])
        wb.save(EXCEL_FILE)
        print(f"✅ Attendance saved for {name}")
    except Exception as e:
        print(f"❌ Excel error: {e}")

def send_email(name, class_name, class_time):
    """Send email notification."""
    if name not in people_emails or name in emails_sent:
        return

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if class_name.startswith("No class"):
        subject = "No Class Today"
        body = f"No class now.\nMarked at {timestamp}"
    else:
        subject = f"Attendance Confirmed: {class_name}"
        body = f"Hello {name},\n\nYour attendance has been marked.\n\n📚 {class_name}\n🕒 {class_time}\n✅ {timestamp}"

    try:
        msg = MIMEText(body)
        msg['Subject'] = subject
        msg['From'] = EMAIL
        msg['To'] = people_emails[name]

        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(EMAIL, PASSWORD)
            server.sendmail(EMAIL, people_emails[name], msg.as_string())
            print(f"📧 Email sent to {name}")
            emails_sent.add(name)
    except Exception as e:
        print(f"❌ Email failed: {e}")

def recognize_face(image_data):
    """Detect & recognize face from base64 image."""
    try:
        img_data = base64.b64decode(image_data.split(",")[1]) if "," in image_data else base64.b64decode(image_data)
    except Exception:
        return {"status": "error", "message": "Invalid image data"}

    frame = cv2.imdecode(np.frombuffer(img_data, np.uint8), cv2.IMREAD_COLOR)
    if frame is None:
        return {"status": "error", "message": "Image decode failed"}

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = detector(gray)
    if not faces:
        return {"status": "no_face", "message": "No face detected"}

    current_class, class_time = get_current_class()

    for face in faces:
        shape = predictor(gray, face)
        encoding = np.array(face_rec_model.compute_face_descriptor(frame, shape))

        distances = [(n, np.linalg.norm(enc - encoding)) for n, enc in face_encodings.items()]
        distances.sort(key=lambda x: x[1])

        if distances and distances[0][1] < 0.5 and (len(distances) == 1 or (distances[1][1] - distances[0][1]) > 0.1):
            name = distances[0][0]
            print(f"😀 Recognized: {name}")
            log_attendance(name, current_class, class_time)
            send_email(name, current_class, class_time)
            return {"status": "recognized", "name": name, "class": current_class, "class_time": class_time}

    os.makedirs("unrecognized_faces", exist_ok=True)
    cv2.imwrite(f"unrecognized_faces/face_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg", frame)
    return {"status": "unrecognized", "message": "Face not matched", "class": current_class, "class_time": class_time}

# ==================== Flask Routes ====================
@app.route("/")
def home():
    return render_template("index.html")

@app.route("/detect", methods=["POST"])
def detect():
    image_data = request.json.get("image")
    if not image_data:
        return jsonify({"status": "error", "message": "No image data"})
    return jsonify(recognize_face(image_data))

@app.route("/download_excel")
def download_excel():
    return send_file(EXCEL_FILE, as_attachment=True) if os.path.exists(EXCEL_FILE) else ("No Excel file found", 404)

# ==================== Run App ====================
if __name__ == "__main__":
    app.run(debug=True)
